import openpyxl
import os

celda = []
fileName = ""


def ReadFile(): # 1 -- read the excel file
    theFile = openpyxl.load_workbook( 'CodesNames.xlsx' )
    allSheetNames = theFile.sheetnames

    print( "===========================================" )
    print( "Extracting data from {}".format( theFile.sheetnames ) )
    print( "===========================================" )

    for sheet in allSheetNames:
        currentSheet = theFile[sheet]
        for row in range( 1, currentSheet.max_row + 1 ):
            for column in "C":  # if add letters, i can add more columns
                C = "{}{}".format( column, row ) # all column A
                ColumnC = currentSheet[C].value

                Code_NAME( ColumnC )

def Code_NAME( codes_names ):
    if codes_names != "":
        celda.append( codes_names )



def loadName():   
    for name in celda:
        dato = name.index('_') # search _
        fileName = name[:dato]+".jpg" # save everythinh before the _   example: 0036
        
        n = name.rstrip() # remove spaces
        newName = n+".jpg" # new name, example 0268_ALFREDO E. ROJAS  CORDERO.jpg
        
        try:
            os.rename(r'C:\\PhotoScan\\Pictures\\'+fileName+'',r'C:\\PhotoScan\\Pictures\\'+newName+'')
        except:
            print("error")
            




ReadFile()
loadName()
